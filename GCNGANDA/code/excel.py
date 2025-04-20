import openpyxl

def create_and_write_excel(file_name="example.xlsx"):
    # 1. 创建一个新的工作簿
    wb = openpyxl.Workbook()

    # 2. 选择默认的活动工作表
    ws = wb.active
    i = 1
    i = str(i)
    # 3. 写入文字
    ws["A"+i] = "Hello"
    ws["B"+i] = "World"
    ws["C"+i] = "这是一个测试"

    # 4. 保存 Excel 文件
    wb.save(file_name)
    print(f"已保存到 {file_name}")

if __name__ == "__main__":
    create_and_write_excel()
